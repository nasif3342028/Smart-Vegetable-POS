import os
import csv
from datetime import datetime
from app_paths import BASE_DIR

try:
    from openpyxl import Workbook, load_workbook  # type: ignore
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side  # type: ignore
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("[EXPORT] openpyxl not installed. Excel export disabled.")


EXPORT_DIR = os.path.join(BASE_DIR, "sales_data")
CSV_FILE = "sales_log.csv"
EXCEL_FILE = "sales_log.xlsx"


class DataExporter:

    def __init__(self, export_dir=EXPORT_DIR):
        self.export_dir = export_dir
        self.csv_path = os.path.join(self.export_dir, CSV_FILE)
        self.excel_path = os.path.join(self.export_dir, EXCEL_FILE)

        if not os.path.exists(self.export_dir):
            os.makedirs(self.export_dir)
            print(f"[EXPORT] Created export folder: {self.export_dir}")

    def export_csv(self, customer_info, cart_items, session_id):
        if not cart_items:
            return

        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        grand_total = sum(item["total"] for item in cart_items)
        file_exists = os.path.exists(self.csv_path)

        try:
            with open(self.csv_path, mode="a", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)

                if not file_exists:
                    writer.writerow([
                        "Session ID", "Customer Name", "Mobile", "Item",
                        "Weight (kg)", "Rate (BDT/kg)", "Item Total (BDT)",
                        "Session Total (BDT)", "Date & Time"
                    ])

                for i, item in enumerate(cart_items):
                    session_total_str = f"{grand_total:.2f}" if i == len(cart_items) - 1 else ""
                    writer.writerow([
                        session_id, customer_info["name"], customer_info["mobile"],
                        item["vegetable"].upper(), f"{item['weight_kg']:.3f}",
                        f"{item['price_per_kg']:.2f}", f"{item['total']:.2f}",
                        session_total_str, now
                    ])

            print(f"[EXPORT] CSV updated: {self.csv_path}")

        except Exception as e:
            print(f"[EXPORT] CSV export failed: {e}")
            raise RuntimeError(f"CSV export failed: {e}")

    def export_excel(self, customer_info, cart_items, session_id):
        if not cart_items:
            return

        if not EXCEL_AVAILABLE:
            print("[EXPORT] Excel export skipped (openpyxl not installed).")
            return

        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        grand_total = sum(item["total"] for item in cart_items)

        header_font = Font(name="Segoe UI", bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="007856", end_color="007856", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        data_font = Font(name="Consolas", size=10)
        data_alignment = Alignment(horizontal="center", vertical="center")

        total_font = Font(name="Segoe UI", bold=True, size=11, color="006600")
        total_fill = PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid")

        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )

        headers = [
            "Session ID", "Customer Name", "Mobile", "Item",
            "Weight (kg)", "Rate (BDT/kg)", "Item Total (BDT)",
            "Session Total (BDT)", "Date & Time"
        ]
        col_widths = [22, 20, 15, 12, 14, 16, 17, 18, 22]

        try:
            if os.path.exists(self.excel_path):
                wb = load_workbook(self.excel_path)
                ws = wb.active
            else:
                wb = Workbook()
                ws = wb.active
                ws.title = "Sales Log"

                ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
                title_cell = ws.cell(row=1, column=1, value="SMART VEGETABLE POS - SALES LOG")
                title_cell.font = Font(name="Segoe UI", bold=True, size=16, color="007856")
                title_cell.alignment = Alignment(horizontal="center", vertical="center")
                ws.row_dimensions[1].height = 30

                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=3, column=col_idx, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = thin_border

                for col_idx, width in enumerate(col_widths, 1):
                    ws.column_dimensions[chr(64 + col_idx)].width = width
                ws.row_dimensions[3].height = 25

            next_row = ws.max_row + 1
            if next_row > 4:
                next_row += 1

            for i, item in enumerate(cart_items):
                row = next_row + i
                session_total_str = f"{grand_total:.2f}" if i == len(cart_items) - 1 else ""
                row_data = [
                    session_id, customer_info["name"], customer_info["mobile"],
                    item["vegetable"].upper(), round(item["weight_kg"], 3),
                    round(item["price_per_kg"], 2), round(item["total"], 2),
                    session_total_str, now
                ]
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row, column=col_idx, value=value)
                    cell.font = data_font
                    cell.alignment = data_alignment
                    cell.border = thin_border

            total_row = next_row + len(cart_items)
            ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=6)
            label_cell = ws.cell(row=total_row, column=1,
                                 value=f"SESSION TOTAL - {customer_info['name'].upper()}")
            label_cell.font = total_font
            label_cell.fill = total_fill
            label_cell.alignment = Alignment(horizontal="right", vertical="center")
            label_cell.border = thin_border

            for col_idx in range(2, 7):
                cell = ws.cell(row=total_row, column=col_idx)
                cell.fill = total_fill
                cell.border = thin_border

            total_cell = ws.cell(row=total_row, column=7, value=grand_total)
            total_cell.font = total_font
            total_cell.fill = total_fill
            total_cell.alignment = Alignment(horizontal="center", vertical="center")
            total_cell.border = thin_border
            total_cell.number_format = "0.00"

            for col_idx in range(8, len(headers) + 1):
                cell = ws.cell(row=total_row, column=col_idx)
                cell.fill = total_fill
                cell.border = thin_border

            wb.save(self.excel_path)
            print(f"[EXPORT] Excel updated: {self.excel_path}")

        except Exception as e:
            print(f"[EXPORT] Excel export failed: {e}")
            raise RuntimeError(f"Excel export failed: {e}")

    def export_all(self, customer_info, cart_items, session_id):
        result = {
            "csv_path": None, "csv_success": False,
            "excel_path": None, "excel_success": False
        }

        try:
            self.export_csv(customer_info, cart_items, session_id)
            result["csv_path"] = self.csv_path
            result["csv_success"] = True
        except RuntimeError as e:
            print(f"[EXPORT] CSV failed: {e}")

        try:
            self.export_excel(customer_info, cart_items, session_id)
            result["excel_path"] = self.excel_path
            result["excel_success"] = True
        except RuntimeError as e:
            print(f"[EXPORT] Excel failed: {e}")

        return result